# ============================================================
#  services/lecturas.py — RF2 / RF3
#  NUEVO: soporte para archivos Excel (.xlsx) del sensor
#  Mantiene compatibilidad total con CSV
# ============================================================

import io, csv, hashlib
from datetime import datetime
from config.db import get_db

# ── Columnas requeridas en CSV ───────────────────────────────
COLUMNAS_CSV = ['fecha_hora','humedad','temperatura','ph',
                'conductividad','salinidad','nitrogeno','fosforo','potasio']

# ── Mapeo de columnas del sensor Excel → nombres internos ───
# Formato: 'nombre_columna_excel': 'nombre_interno'
COLUMNAS_EXCEL = {
    'time':               'fecha_hora',
    'temp(℃)':            'temperatura',
    'temp(c)':            'temperatura',   # fallback sin símbolo
    'hum(%)':             'humedad',
    'conductivity(us/cm)':'conductividad',
    'ph':                 'ph',
    'n(mg/kg)':           'nitrogeno',
    'p(mg/kg)':           'fosforo',
    'k(mg/kg)':           'potasio',
}

RANGOS = {
    'humedad':       (0,   100),
    'temperatura':   (-10,  60),
    'ph':            (0,    14),
    'conductividad': (0,  9999),
    'salinidad':     (0,    50),
    'nitrogeno':     (0,   500),
    'fosforo':       (0,   500),
    'potasio':       (0,  2000),
}


# ── Hash MD5 ─────────────────────────────────────────────────
def calcular_hash(contenido_bytes: bytes) -> str:
    return hashlib.md5(contenido_bytes).hexdigest()


# ── Verificar duplicado ──────────────────────────────────────
def verificar_duplicado(hash_md5: str, area_id: int):
    with get_db() as (conn, cursor):
        cursor.execute("""
            SELECT c.CargaID, c.NombreArchivo, c.FechaCarga, u.Nombre AS Usuario
            FROM dbo.CargasCSV c
            JOIN dbo.Usuarios u ON u.UsuarioID = c.UsuarioID
            WHERE c.HashMD5 = ? AND c.AreaID = ? AND c.Estado = 'Procesado'
        """, hash_md5, area_id)
        row = cursor.fetchone()
        if row:
            return {
                "cargaId":       row.CargaID,
                "nombreArchivo": row.NombreArchivo,
                "fechaCarga":    str(row.FechaCarga)[:19],
                "usuario":       row.Usuario,
            }
        return None


# ── Punto de entrada principal ───────────────────────────────
def procesar_csv(area_id: int, usuario_id: int, nombre_archivo: str,
                 contenido_bytes: bytes) -> dict:
    """
    Acepta tanto archivos CSV como Excel (.xlsx / .xls).
    Detecta el tipo por la extensión del nombre de archivo.
    """
    hash_md5 = calcular_hash(contenido_bytes)

    # Bloquear duplicado
    dup = verificar_duplicado(hash_md5, area_id)
    if dup:
        raise ValueError(
            f"DUPLICADO|{dup['nombreArchivo']}|{dup['fechaCarga']}|{dup['usuario']}"
        )

    ext = nombre_archivo.lower().split('.')[-1]

    if ext in ('xlsx', 'xls', 'xlsm'):
        filas_ok = _leer_excel(contenido_bytes, ext)
    else:
        filas_ok = _leer_csv(contenido_bytes)

    if not filas_ok:
        raise ValueError("El archivo no contiene filas de datos válidas.")

    # Una sola transacción
    with get_db() as (conn, cursor):
        cursor.execute("""
            INSERT INTO dbo.CargasCSV
                (AreaID, UsuarioID, NombreArchivo, TotalFilas, Estado, HashMD5)
            OUTPUT INSERTED.CargaID
            VALUES (?,?,?,?,'Pendiente',?)
        """, area_id, usuario_id, nombre_archivo, len(filas_ok), hash_md5)
        carga_id = cursor.fetchone()[0]

        for l in filas_ok:
            cursor.execute("""
                INSERT INTO dbo.LecturasSensor
                    (AreaID, CargaID, FechaHoraMedicion,
                     Humedad_Pct, Temperatura_C, pH,
                     ConductividadElectrica, Salinidad_Ppt,
                     Nitrogeno_ppm, Fosforo_ppm, Potasio_ppm)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, area_id, carga_id, l['fecha_hora'],
                l['humedad'], l['temperatura'], l['ph'],
                l['conductividad'], l['salinidad'],
                l['nitrogeno'], l['fosforo'], l['potasio'])

        cursor.execute("""
            UPDATE dbo.CargasCSV
            SET FilasProcesadas=?, Estado='Procesado' WHERE CargaID=?
        """, len(filas_ok), carga_id)

    return {
        "cargaId":         carga_id,
        "totalFilas":      len(filas_ok),
        "filasInsertadas": len(filas_ok),
        "estado":          "Procesado",
        "hashMD5":         hash_md5,
    }


# ── Lector de Excel ──────────────────────────────────────────
def _leer_excel(contenido_bytes: bytes, ext: str) -> list:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("openpyxl no está instalado. Ejecuta: pip install openpyxl")

    wb = load_workbook(io.BytesIO(contenido_bytes), read_only=True, data_only=True)
    ws = wb.active

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        raise ValueError("El archivo Excel está vacío.")

    # Normalizar encabezados de la primera fila
    encabezados_raw = [str(c).strip().lower() if c else '' for c in filas[0]]

    # Mapear columnas del sensor a nombres internos
    mapa_idx = {}  # nombre_interno -> índice de columna
    for idx, enc in enumerate(encabezados_raw):
        nombre_interno = COLUMNAS_EXCEL.get(enc)
        if nombre_interno:
            mapa_idx[nombre_interno] = idx

    # Verificar que están las columnas mínimas (salinidad es opcional en Excel)
    requeridas_excel = ['fecha_hora','temperatura','humedad','conductividad','ph','fosforo','potasio']
    faltantes = [c for c in requeridas_excel if c not in mapa_idx]
    if faltantes:
        raise ValueError(f"Columnas no encontradas en el Excel: {', '.join(faltantes)}")

    filas_ok = []
    errores  = []
    for num, fila in enumerate(filas[1:], start=2):
        if not fila or all(v is None for v in fila):
            continue  # saltar filas vacías
        try:
            filas_ok.append(_parsear_fila_excel(fila, mapa_idx, num))
        except ValueError as e:
            errores.append(str(e))
            if len(errores) >= 3:
                errores.append("(se omiten errores adicionales...)")
                break

    if errores:
        raise ValueError("Errores en el Excel: " + " | ".join(errores))

    return filas_ok


def _parsear_fila_excel(fila: tuple, mapa_idx: dict, num: int) -> dict:
    def _val(key, requerido=True):
        idx = mapa_idx.get(key)
        if idx is None:
            return 0.0  # columna opcional ausente → 0
        raw = fila[idx] if idx < len(fila) else None
        if raw is None or str(raw).strip() == '':
            if requerido:
                raise ValueError(f"Fila {num}: valor vacío en '{key}'")
            return 0.0
        try:
            return float(str(raw).replace(',', '.'))
        except ValueError:
            raise ValueError(f"Fila {num}: '{raw}' no es número en '{key}'")

    # Parsear fecha
    idx_fecha = mapa_idx.get('fecha_hora')
    raw_fecha = fila[idx_fecha] if idx_fecha is not None and idx_fecha < len(fila) else None
    fecha = _parsear_fecha(raw_fecha, num)

    l = {
        'fecha_hora':    fecha,
        'humedad':       _val('humedad'),
        'temperatura':   _val('temperatura'),
        'ph':            _val('ph'),
        'conductividad': _val('conductividad'),
        'salinidad':     _val('salinidad', requerido=False),  # Excel no trae salinidad
        'nitrogeno':     _val('nitrogeno', requerido=False),
        'fosforo':       _val('fosforo'),
        'potasio':       _val('potasio'),
    }

    # Validar rangos (con tolerancia para nitrógeno que puede ser 0)
    rangos_validar = {k: v for k, v in RANGOS.items() if k != 'nitrogeno'}
    for campo, (mn, mx) in rangos_validar.items():
        val = l[campo]
        if not (mn <= val <= mx):
            raise ValueError(f"Fila {num}: {campo}={val} fuera de rango [{mn},{mx}]")
    return l


# ── Lector de CSV ────────────────────────────────────────────
def _leer_csv(contenido_bytes: bytes) -> list:
    try:
        texto = contenido_bytes.decode('utf-8-sig')
    except UnicodeDecodeError:
        texto = contenido_bytes.decode('latin-1')

    reader = csv.DictReader(io.StringIO(texto))
    if not reader.fieldnames:
        raise ValueError("El archivo CSV está vacío o no tiene encabezados.")

    reader.fieldnames = [
        c.strip().lower().replace(' ','_').replace('"','').replace("'",'')
         .replace('é','e').replace('ó','o').replace('ú','u')
         .replace('á','a').replace('í','i')
        for c in reader.fieldnames
    ]

    faltantes = [c for c in COLUMNAS_CSV if c not in reader.fieldnames]
    if faltantes:
        raise ValueError(f"Columnas faltantes en CSV: {', '.join(faltantes)}")

    filas_raw = list(reader)
    if not filas_raw:
        raise ValueError("El CSV no contiene filas de datos.")

    filas_ok, errores = [], []
    for i, fila in enumerate(filas_raw, start=2):
        try:
            filas_ok.append(_parsear_fila_csv(fila, i))
        except ValueError as e:
            errores.append(str(e))
            if len(errores) >= 3:
                errores.append("(se omiten errores adicionales...)")
                break

    if errores:
        raise ValueError("Errores en CSV: " + " | ".join(errores))
    return filas_ok


def _parsear_fila_csv(fila: dict, num: int) -> dict:
    def _num(key):
        val = fila.get(key,'').strip()
        if not val: raise ValueError(f"Fila {num}: vacío en '{key}'")
        try: return float(val.replace(',','.'))
        except ValueError: raise ValueError(f"Fila {num}: '{val}' no es número en '{key}'")

    fecha = _parsear_fecha(fila.get('fecha_hora','').strip(), num)
    l = {
        'fecha_hora':    fecha,
        'humedad':       _num('humedad'),
        'temperatura':   _num('temperatura'),
        'ph':            _num('ph'),
        'conductividad': _num('conductividad'),
        'salinidad':     _num('salinidad'),
        'nitrogeno':     _num('nitrogeno'),
        'fosforo':       _num('fosforo'),
        'potasio':       _num('potasio'),
    }
    for campo, (mn, mx) in RANGOS.items():
        if not (mn <= l[campo] <= mx):
            raise ValueError(f"Fila {num}: {campo}={l[campo]} fuera de rango [{mn},{mx}]")
    return l


# ── Parsear fecha (compartido) ───────────────────────────────
def _parsear_fecha(raw, num: int) -> datetime:
    if raw is None or str(raw).strip() == '':
        raise ValueError(f"Fila {num}: fecha vacía")

    # Si ya es datetime (openpyxl puede devolver datetime directamente)
    if isinstance(raw, datetime):
        return raw

    raw = str(raw).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S',
                '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M',
                '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    raise ValueError(f"Fila {num}: formato de fecha no reconocido '{raw}'")


# ── Consultas ────────────────────────────────────────────────
def obtener_lecturas(area_id: int, desde: str, hasta: str) -> list:
    with get_db() as (conn, cursor):
        cursor.execute("""
            SELECT FechaHoraMedicion, Humedad_Pct, Temperatura_C, pH,
                   ConductividadElectrica, Salinidad_Ppt,
                   Nitrogeno_ppm, Fosforo_ppm, Potasio_ppm
            FROM dbo.LecturasSensor
            WHERE AreaID=? AND FechaHoraMedicion BETWEEN ? AND ?
            ORDER BY FechaHoraMedicion
        """, area_id, desde, hasta)
        return [{"fechaHora":str(r.FechaHoraMedicion),"humedad":_f(r.Humedad_Pct),
                 "temperatura":_f(r.Temperatura_C),"pH":_f(r.pH),
                 "conductividad":_f(r.ConductividadElectrica),"salinidad":_f(r.Salinidad_Ppt),
                 "nitrogeno":_f(r.Nitrogeno_ppm),"fosforo":_f(r.Fosforo_ppm),
                 "potasio":_f(r.Potasio_ppm)} for r in cursor.fetchall()]


def obtener_estadisticos(area_id: int, desde: str, hasta: str) -> dict:
    with get_db() as (conn, cursor):
        cursor.execute("""
            SELECT COUNT(*) AS TotalLecturas,
                ROUND(AVG(Humedad_Pct),2) AS Humedad_Prom, ROUND(MAX(Humedad_Pct),2) AS Humedad_Max, ROUND(MIN(Humedad_Pct),2) AS Humedad_Min,
                ROUND(AVG(Temperatura_C),2) AS Temp_Prom,  ROUND(MAX(Temperatura_C),2) AS Temp_Max,  ROUND(MIN(Temperatura_C),2) AS Temp_Min,
                ROUND(AVG(pH),2) AS pH_Prom,               ROUND(MAX(pH),2) AS pH_Max,               ROUND(MIN(pH),2) AS pH_Min,
                ROUND(AVG(ConductividadElectrica),2) AS CE_Prom, ROUND(MAX(ConductividadElectrica),2) AS CE_Max, ROUND(MIN(ConductividadElectrica),2) AS CE_Min,
                ROUND(AVG(Salinidad_Ppt),3) AS Sal_Prom,   ROUND(MAX(Salinidad_Ppt),3) AS Sal_Max,   ROUND(MIN(Salinidad_Ppt),3) AS Sal_Min,
                ROUND(AVG(Nitrogeno_ppm),2) AS N_Prom,     ROUND(MAX(Nitrogeno_ppm),2) AS N_Max,     ROUND(MIN(Nitrogeno_ppm),2) AS N_Min,
                ROUND(AVG(Fosforo_ppm),2) AS P_Prom,       ROUND(MAX(Fosforo_ppm),2) AS P_Max,       ROUND(MIN(Fosforo_ppm),2) AS P_Min,
                ROUND(AVG(Potasio_ppm),2) AS K_Prom,       ROUND(MAX(Potasio_ppm),2) AS K_Max,       ROUND(MIN(Potasio_ppm),2) AS K_Min
            FROM dbo.LecturasSensor WHERE AreaID=? AND FechaHoraMedicion BETWEEN ? AND ?
        """, area_id, desde, hasta)
        r = cursor.fetchone()
        if not r or r.TotalLecturas == 0: return {}
        return {d[0]: _f(getattr(r, d[0])) for d in cursor.description}


def historial_cargas(area_id: int) -> list:
    with get_db() as (conn, cursor):
        cursor.execute("""
            SELECT c.CargaID, c.NombreArchivo, c.FechaCarga,
                   c.TotalFilas, c.FilasProcesadas, c.Estado, u.Nombre AS Usuario
            FROM dbo.CargasCSV c
            JOIN dbo.Usuarios u ON u.UsuarioID = c.UsuarioID
            WHERE c.AreaID=? ORDER BY c.FechaCarga DESC
        """, area_id)
        return [{"cargaId":r.CargaID,"nombreArchivo":r.NombreArchivo,
                 "fechaCarga":str(r.FechaCarga),"totalFilas":r.TotalFilas,
                 "filasProcesadas":r.FilasProcesadas,"estado":r.Estado,
                 "usuario":r.Usuario} for r in cursor.fetchall()]


def exportar_csv(area_id: int, desde: str, hasta: str) -> str:
    lecturas = obtener_lecturas(area_id, desde, hasta)
    enc = "timestamp,humedad,temperatura,pH,conductividad,salinidad,nitrogeno,fosforo,potasio"
    if not lecturas: return enc + "\n"
    return enc + '\n' + '\n'.join(
        f"{l['fechaHora']},{l['humedad']},{l['temperatura']},{l['pH']},"
        f"{l['conductividad']},{l['salinidad']},{l['nitrogeno']},{l['fosforo']},{l['potasio']}"
        for l in lecturas)


def _f(val): return float(val) if val is not None else None